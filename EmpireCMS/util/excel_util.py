import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font,Side,Border
from util.datetime_util import *
from util.global_var import *
from util.log_util import *

#支持excel读写操作的工具类
class Excel:
    # 初始化读取excel文件
    def __init__(self, file_path):
        if not os.path.exists(file_path):
            error("Excel工具类初始化失败：【{}】文件不存在！".format(file_path))
            return
        if not file_path.endswith(".xlsx") or not file_path.endswith(".xlsx"):
            error("Excel工具类初始化失败：【{}】文件非excel文件类型！".format(file_path))
            return

        self.wb = load_workbook(file_path)
        # 初始化默认sheet
        self.ws = self.wb.active
        self.data_file_path = file_path
        # 初始化颜色字典，共设置样式
        self.color_dict = {"red": "FFFF3030", "green": "FF008B00"}

    # 获取所有sheet
    def get_all_sheet(self):
        # 使用注释代码会报错
        # return self.wb.get_sheet_names()
        return self.wb.sheetnames

    def get_sheet(self, sheet_name):
        '''
        打开指定sheet
        :param sheet_name: 表名
        '''
        if sheet_name not in self.get_all_sheet():
            error("sheet名称【%s】不存在！" % sheet_name)
            return
        # self.ws = self.wb.get_sheet_by_name(sheet_name)
        self.ws = self.wb[sheet_name]
        return True

    def get_max_row_no(self):
        '''
        获取最大行号
        openpyxl的API的行、列索引默认都从1开始
        '''
        return self.ws.max_row

    def get_max_col_no(self):
        '''
        获取最大列号
        '''
        return self.ws.max_column

    def get_all_row_data(self, head_line=True):
        '''
        获取所有行数据
        :param head_line:
        '''
        # 是否需要标题行数据的标识，默认需要
        if head_line:
            min_row = 1  # 行号从1开始，即1为标题
        else:
            min_row = 2
        result = []
        # min_row=None：默认获取标题行数据
        for row in self.ws.iter_rows(min_row=min_row, max_row=self.get_max_row_no(), max_col=self.get_max_col_no()):
            result.append([cell.value for cell in row])
        return result

    def get_row_data(self, row_num):
        '''
        获取指定行数据
        :param row_num: 行号
        '''
        # 0 为标题行
        return [cell.value for cell in self.ws[row_num + 1]]

    def get_col_data(self, col_num):
        '''
        获取指定列数据
        :param col_num: 列号
        '''
        # 索引从0开始
        return [cell.value for cell in tuple(self.ws.columns)[col_num]]

    def write_row_data(self, data, font_color=None, border=True, fill_color=None):
        '''
        追加行数据且可以设置样式
        :param data:内容
        :param font_color:字体颜色
        :param border:边框
        :param fill_color:填充颜色
        '''
        if not isinstance(data, (list, tuple)):
            print("写入数据失败：数据不为列表或元组类型！【%s】" % data)
        self.ws.append(data)
        # 设置字体颜色
        if font_color:
            if font_color.lower() in self.color_dict.keys():
                font_color = self.color_dict[font_color]
        # 设置单元格填充颜色
        if fill_color:
            if fill_color.lower() in self.color_dict.keys():
                fill_color = self.color_dict[fill_color]
        # 设置单元格边框
        if border:
            bd = Side(style="thin", color="000000")

        # 记录数据长度（否则会默认与之前行最长数据行的长度相同，导致样式超过了该行实际长度）
        count = 0
        for cell in self.ws[self.get_max_row_no()]:
            # 设置完该行的实际数据长度样式后，则退出
            if count > len(data) - 1:
                break
            if font_color:
                cell.font = Font(color=font_color)
            # 如果没有设置字体颜色，则默认给执行结果添加字体颜色
            else:
                if cell.value is not None and isinstance(cell.value, str):
                    if cell.value.lower() == "pass" or cell.value == "成功":
                        cell.font = Font(color=self.color_dict["green"])
                    elif cell.value.lower() == "fail" or cell.value == "失败":
                        cell.font = Font(color=self.color_dict["red"])
            if border:
                cell.border = Border(left=bd, right=bd, top=bd, bottom=bd)
            if fill_color:
                cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
            count += 1

    def insert_row_data(self, row_no, data, font_color=None, border=True, fill_color=None):
        '''
        指定行插入数据（行索引从0开始）
        :param row_no:
        :param data:
        :param font_color:
        :param border:
        :param fill_color:
        '''
        if not isinstance(data, (list, tuple)):
            print("写入数据失败：数据不为列表或元组类型！【%s】" % data)
        for idx, cell in enumerate(self.ws[row_no + 1]):  # 此处行索引从1开始
            cell.value = data[idx]

    def save(self, save_file_name, timestamp):
        '''
        生成写入了测试结果的excel数据文件
        :param save_file_name:
        :param timestamp:
        :return:
        '''
        save_dir = os.path.join(TEST_REPORT_FILE_DIR, get_chinese_date())
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        save_file = os.path.join(save_dir, save_file_name + '' + timestamp + ".xlsx")
        self.wb.save(save_file)
        info("生成测试结果文件：%s" % save_file)
        return save_file
if __name__ == "__main__":
    from util.global_var import *
    from util.datetime_util import *
    excel = Excel(TEST_DATA_FILE_PATH)
    print(excel.get_all_sheet())
    print(excel.get_sheet("登录（非数据驱动）"))
    excel.get_sheet("测试结果统计")
    excel.write_row_data(["4", None, "嘻哈"], "green", True, "red")
    # excel.insert_row_data(1, [1, 2, 3])
    excel.save("data", get_timestamp())